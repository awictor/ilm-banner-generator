"""
"Just Dropped" Instagram Story Franchise Generator

4-step wizard with human approval gates:
  Step 1: Theme Development (automated research + approval)
  Step 2: Product Curation (manual ASIN entry)
  Step 3: Creative Generation (5 channel-specific engines)
  Step 4: Reporting (Excel export)
"""

import zipfile
from io import BytesIO

import streamlit as st
from PIL import Image

import story_engine
import story_themes
from shared import image_picker, fetch_image_from_url, remove_background

# ── Channel definitions ──────────────────────────────────────────
CHANNELS = [
    "@AmazonHome",
    "@AmazonBeauty",
    "@AmazonFashion",
    "@Amazon",
    "@Amazon.ca",
]

US_CHANNELS = ["@AmazonHome", "@AmazonBeauty", "@AmazonFashion", "@Amazon"]

# ── Session state initialization ─────────────────────────────────
def _init_state():
    defaults = {
        "jd_step": 1,
        "jd_themes": None,           # {channel: [theme_dicts]}
        "jd_approved_themes": {},     # {channel: theme_name}
        "jd_products": {},            # {channel: [product_dicts]}
        "jd_alt_asins": {},           # {channel: [alt_asin_dicts]}
        "jd_generated_frames": None,  # [(filename, BytesIO)]
        "jd_report_data": None,       # for Excel export
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init_state()


# ── Navigation helpers ───────────────────────────────────────────
def _go_to_step(step):
    st.session_state.jd_step = step


# ── Page header ──────────────────────────────────────────────────
st.title("Just Dropped — Story Generator")

# Step indicator
cols = st.columns(4)
step_labels = ["1. Theme", "2. Products", "3. Generate", "4. Report"]
for i, (col, label) in enumerate(zip(cols, step_labels), start=1):
    with col:
        if i == st.session_state.jd_step:
            st.markdown(f"**:blue[{label}]**")
        elif i < st.session_state.jd_step:
            st.markdown(f"~~{label}~~ :white_check_mark:")
        else:
            st.markdown(f"*{label}*")

st.divider()


# ══════════════════════════════════════════════════════════════════
# STEP 1: Theme Development
# ══════════════════════════════════════════════════════════════════
if st.session_state.jd_step == 1:
    st.header("Step 1: Theme Development")
    st.markdown(
        "Research current trends and select a theme for each channel's franchise."
    )

    # Research button
    if st.button("Research Trends", type="primary"):
        with st.spinner("Researching trends via Brave Search API..."):
            themes = story_themes.research_themes()
        st.session_state.jd_themes = themes
        st.rerun()

    # Show theme proposals and let user approve
    if st.session_state.jd_themes:
        approved = st.session_state.jd_approved_themes

        for channel in CHANNELS:
            channel_themes = st.session_state.jd_themes.get(channel, [])
            if not channel_themes:
                continue

            st.subheader(channel)
            options = [t["name"] for t in channel_themes]
            descriptions = {t["name"]: t["rationale"] for t in channel_themes}

            selected = st.radio(
                f"Select theme for {channel}",
                options=options,
                key=f"theme_select_{channel}",
                index=options.index(approved[channel]) if channel in approved else 0,
            )

            # Show rationale for selected theme
            if selected and selected in descriptions:
                st.info(descriptions[selected])

            approved[channel] = selected

        st.session_state.jd_approved_themes = approved

        # Allow custom theme entry
        st.markdown("---")
        st.markdown("**Or enter a custom theme for any channel:**")
        for channel in CHANNELS:
            custom = st.text_input(
                f"Custom theme for {channel} (leave blank to use selection above)",
                key=f"custom_theme_{channel}",
            )
            if custom:
                st.session_state.jd_approved_themes[channel] = custom

        st.divider()

        # Check all channels have approved themes
        all_approved = all(ch in st.session_state.jd_approved_themes for ch in CHANNELS)
        if st.button("Approve Themes & Continue", type="primary", disabled=not all_approved):
            _go_to_step(2)
            st.rerun()

        if not all_approved:
            st.warning("Select a theme for each channel to continue.")


# ══════════════════════════════════════════════════════════════════
# STEP 2: Product Curation
# ══════════════════════════════════════════════════════════════════
elif st.session_state.jd_step == 2:
    st.header("Step 2: Product Curation")
    st.markdown(
        "Enter 10 ASINs per franchise (50 total). "
        "For each ASIN, provide brand name, product name, product copy, and an image."
    )

    # Back button
    if st.button("← Back to Theme Selection"):
        _go_to_step(1)
        st.rerun()

    # Tab per channel
    tabs = st.tabs(CHANNELS)

    for tab, channel in zip(tabs, CHANNELS):
        with tab:
            st.subheader(f"{channel} — Primary ASINs (10)")

            if channel not in st.session_state.jd_products:
                st.session_state.jd_products[channel] = [
                    {"asin": "", "brand": "", "product_name": "", "copy": "", "image": None}
                    for _ in range(10)
                ]

            products = st.session_state.jd_products[channel]

            for i in range(10):
                with st.expander(f"Product {i+1}" + (f" — {products[i]['asin']}" if products[i]['asin'] else ""), expanded=(i == 0)):
                    c1, c2 = st.columns([1, 1])
                    with c1:
                        products[i]["asin"] = st.text_input(
                            "ASIN", value=products[i]["asin"],
                            key=f"asin_{channel}_{i}",
                            placeholder="e.g. B0XXXXXXXXX"
                        )
                        products[i]["brand"] = st.text_input(
                            "Brand Name", value=products[i]["brand"],
                            key=f"brand_{channel}_{i}"
                        )
                        products[i]["product_name"] = st.text_input(
                            "Product Name", value=products[i]["product_name"],
                            key=f"prodname_{channel}_{i}"
                        )
                        products[i]["copy"] = st.text_area(
                            "Product Copy", value=products[i]["copy"],
                            key=f"copy_{channel}_{i}",
                            height=80
                        )

                    with c2:
                        st.markdown("**Product Image**")
                        img = image_picker(
                            f"product image for {channel} #{i+1}",
                            f"jd_{channel}_{i}"
                        )
                        if img:
                            products[i]["image"] = img

            st.session_state.jd_products[channel] = products

            # Alternative ASINs section
            st.markdown("---")
            st.subheader(f"{channel} — Alternative ASINs (40)")
            st.markdown("Lighter-weight: just ASIN + Brand")

            if channel not in st.session_state.jd_alt_asins:
                st.session_state.jd_alt_asins[channel] = [
                    {"asin": "", "brand": ""} for _ in range(40)
                ]

            alt_asins = st.session_state.jd_alt_asins[channel]

            # Show in a compact grid (4 columns)
            for row_start in range(0, 40, 4):
                row_cols = st.columns(4)
                for col_idx, col in enumerate(row_cols):
                    idx = row_start + col_idx
                    if idx < 40:
                        with col:
                            alt_asins[idx]["asin"] = st.text_input(
                                f"ASIN {idx+1}", value=alt_asins[idx]["asin"],
                                key=f"alt_asin_{channel}_{idx}",
                                label_visibility="collapsed",
                                placeholder=f"Alt ASIN {idx+1}"
                            )
                            alt_asins[idx]["brand"] = st.text_input(
                                f"Brand {idx+1}", value=alt_asins[idx]["brand"],
                                key=f"alt_brand_{channel}_{idx}",
                                label_visibility="collapsed",
                                placeholder=f"Brand {idx+1}"
                            )

            st.session_state.jd_alt_asins[channel] = alt_asins

    # Validation
    st.divider()
    st.subheader("Validation")

    validation_errors = []
    validation_warnings = []

    # Check uniqueness within each franchise
    for channel in CHANNELS:
        products = st.session_state.jd_products.get(channel, [])
        asins = [p["asin"] for p in products if p["asin"]]
        dupes = set(a for a in asins if asins.count(a) > 1)
        if dupes:
            validation_errors.append(
                f"{channel}: Duplicate ASINs within franchise: {', '.join(dupes)}"
            )

    # Check brand overlaps between US franchises
    us_brand_map = {}  # brand -> list of channels
    for channel in US_CHANNELS:
        products = st.session_state.jd_products.get(channel, [])
        for p in products:
            if p["brand"]:
                brand = p["brand"].strip().lower()
                us_brand_map.setdefault(brand, []).append(channel)

    for brand, channels_with in us_brand_map.items():
        unique_channels = list(set(channels_with))
        if len(unique_channels) > 1:
            validation_warnings.append(
                f'Brand "{brand}" appears in multiple US franchises: {", ".join(unique_channels)}'
            )

    # Check that all products have images
    missing_images = []
    for channel in CHANNELS:
        products = st.session_state.jd_products.get(channel, [])
        for i, p in enumerate(products):
            if p["asin"] and p["image"] is None:
                missing_images.append(f"{channel} Product {i+1} ({p['asin']})")

    if missing_images:
        validation_errors.append(
            f"Missing images for: {', '.join(missing_images[:5])}"
            + (f" and {len(missing_images)-5} more" if len(missing_images) > 5 else "")
        )

    if validation_errors:
        for err in validation_errors:
            st.error(err)
    if validation_warnings:
        for warn in validation_warnings:
            st.warning(warn)
    if not validation_errors and not validation_warnings:
        st.success("All validations passed.")

    # Count entered products
    total_entered = sum(
        1 for ch in CHANNELS
        for p in st.session_state.jd_products.get(ch, [])
        if p["asin"] and p["image"] is not None
    )
    st.metric("Products with image ready", f"{total_entered} / 50")

    can_proceed = total_entered > 0 and not validation_errors

    if st.button("Continue to Generation", type="primary", disabled=not can_proceed):
        _go_to_step(3)
        st.rerun()


# ══════════════════════════════════════════════════════════════════
# STEP 3: Creative Generation
# ══════════════════════════════════════════════════════════════════
elif st.session_state.jd_step == 3:
    st.header("Step 3: Creative Generation")
    st.markdown("Generate 1080×1920px Instagram Story frames for all 5 channels.")

    if st.button("← Back to Product Curation"):
        _go_to_step(2)
        st.rerun()

    # Summary of what will be generated
    st.subheader("Generation Summary")
    for channel in CHANNELS:
        products = st.session_state.jd_products.get(channel, [])
        ready = [p for p in products if p["asin"] and p["image"] is not None]
        n = len(ready)
        frame_count = n + 2 if n > 0 else 0  # collage + individuals + collage dupe
        extra = ""
        if channel == "@Amazon.ca" and n > 0:
            frame_count *= 2  # English + French
            extra = " (EN + FR)"
        st.markdown(f"- **{channel}**: {n} products → {frame_count} frames{extra}")

    st.divider()

    # Generate button
    if st.button("Generate All Frames", type="primary"):
        franchise_data = {}
        for channel in CHANNELS:
            products = st.session_state.jd_products.get(channel, [])
            ready = [p for p in products if p["asin"] and p["image"] is not None]
            if ready:
                franchise_data[channel] = ready

        theme_names = st.session_state.jd_approved_themes

        with st.spinner("Generating frames... This may take a moment."):
            all_frames = story_engine.generate_all_franchises(franchise_data, theme_names)

        st.session_state.jd_generated_frames = all_frames
        st.rerun()

    # Show generated frames
    if st.session_state.jd_generated_frames:
        frames = st.session_state.jd_generated_frames
        st.success(f"Generated {len(frames)} frames!")

        # Group by folder
        folders = {}
        for fname, buf in frames:
            folder = fname.split("/")[0] if "/" in fname else "Other"
            folders.setdefault(folder, []).append((fname, buf))

        # Preview by folder
        for folder in sorted(folders.keys()):
            with st.expander(f"{folder} ({len(folders[folder])} frames)", expanded=False):
                for fname, buf in folders[folder]:
                    st.caption(fname)
                    buf.seek(0)
                    st.image(buf, width=270)

        # Download ZIP
        st.divider()
        zip_buf = BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for fname, buf in frames:
                buf.seek(0)
                zf.writestr(fname, buf.read())
        zip_buf.seek(0)

        st.download_button(
            label="Download All Frames (ZIP)",
            data=zip_buf,
            file_name="Just_Dropped_Frames.zip",
            mime="application/zip",
            type="primary",
        )

        st.divider()
        if st.button("Continue to Report", type="primary"):
            _go_to_step(4)
            st.rerun()


# ══════════════════════════════════════════════════════════════════
# STEP 4: Reporting (Excel Export)
# ══════════════════════════════════════════════════════════════════
elif st.session_state.jd_step == 4:
    st.header("Step 4: Reporting")
    st.markdown("Generate an Excel report with one sheet per franchise.")

    if st.button("← Back to Generation"):
        _go_to_step(3)
        st.rerun()

    if st.button("Generate Excel Report", type="primary"):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment

            wb = openpyxl.Workbook()
            # Remove default sheet
            wb.remove(wb.active)

            for channel in CHANNELS:
                products = st.session_state.jd_products.get(channel, [])
                ready = [p for p in products if p["asin"]]
                if not ready:
                    continue

                safe_name = channel.replace("@", "").replace(".", "_")[:31]
                ws = wb.create_sheet(title=safe_name)

                # Headers
                headers = [
                    "Frame #",
                    "Child ASIN",
                    "Brand",
                    "Product Name",
                    "MCID",
                    "ASIN Site Launch Date",
                    "WBR Team",
                    "Account Rep",
                ]
                header_font = Font(bold=True)
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                # Data rows – mirror frame sequence
                # Frame 1 is collage (no single ASIN)
                ws.cell(row=2, column=1, value="Frame 1 (Collage)")
                ws.cell(row=2, column=2, value="—")

                for i, prod in enumerate(ready):
                    row = i + 3  # row 3 onwards
                    ws.cell(row=row, column=1, value=f"Frame {i+2}")
                    ws.cell(row=row, column=2, value=prod["asin"])
                    ws.cell(row=row, column=3, value=prod.get("brand", ""))
                    ws.cell(row=row, column=4, value=prod.get("product_name", ""))
                    # MCID, Launch Date, WBR Team, Account Rep left blank for manual fill

                # Last frame is collage duplicate
                last_row = len(ready) + 3
                ws.cell(row=last_row, column=1, value=f"Frame {len(ready)+2} (Collage)")
                ws.cell(row=last_row, column=2, value="—")

                # Alternative ASINs section
                alt_row = last_row + 2
                ws.cell(row=alt_row, column=1, value="Alternative ASINs")
                ws.cell(row=alt_row, column=1).font = Font(bold=True)

                alt_headers = ["#", "ASIN", "Brand"]
                for col_idx, h in enumerate(alt_headers, start=1):
                    cell = ws.cell(row=alt_row + 1, column=col_idx, value=h)
                    cell.font = Font(bold=True)

                alt_asins = st.session_state.jd_alt_asins.get(channel, [])
                for j, alt in enumerate(alt_asins):
                    if alt["asin"]:
                        r = alt_row + 2 + j
                        ws.cell(row=r, column=1, value=j + 1)
                        ws.cell(row=r, column=2, value=alt["asin"])
                        ws.cell(row=r, column=3, value=alt.get("brand", ""))

                # Auto-size columns
                for col in ws.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

            # Save to buffer
            excel_buf = BytesIO()
            wb.save(excel_buf)
            excel_buf.seek(0)

            st.session_state.jd_report_data = excel_buf
            st.success("Excel report generated!")

        except ImportError:
            st.error("openpyxl is required for Excel export. Install it with: pip install openpyxl")

    # Download button
    if st.session_state.jd_report_data:
        st.session_state.jd_report_data.seek(0)
        st.download_button(
            label="Download Excel Report",
            data=st.session_state.jd_report_data,
            file_name="Just_Dropped_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )

    # Summary
    st.divider()
    st.subheader("Summary")

    total_frames = len(st.session_state.jd_generated_frames or [])
    total_products = sum(
        1 for ch in CHANNELS
        for p in st.session_state.jd_products.get(ch, [])
        if p["asin"]
    )
    st.metric("Total Frames Generated", total_frames)
    st.metric("Total Products", total_products)

    st.markdown("**Approved Themes:**")
    for ch, theme in st.session_state.jd_approved_themes.items():
        st.markdown(f"- {ch}: *{theme}*")

    # Start over
    st.divider()
    if st.button("Start New Franchise"):
        for key in list(st.session_state.keys()):
            if key.startswith("jd_"):
                del st.session_state[key]
        st.rerun()
